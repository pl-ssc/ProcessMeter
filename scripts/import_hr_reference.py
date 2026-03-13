#!/usr/bin/env python3
import argparse
import json
import os
import sys
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


def parse_args():
    parser = argparse.ArgumentParser(
        description="Normalize HR process hierarchy from XLSX and emit SQL for refdb import."
    )
    parser.add_argument(
        "--input",
        default="/Users/romangaleev/Desktop/20230124 HR_X_v4.xlsx",
        help="Path to the source XLSX file.",
    )
    parser.add_argument(
        "--output-sql",
        default="tmp/hr_reference_import.sql",
        help="Path to the generated SQL file.",
    )
    parser.add_argument(
        "--output-json",
        default="tmp/hr_reference_structure.json",
        help="Path to the generated normalized JSON file.",
    )
    parser.add_argument(
        "--process-1-id",
        type=int,
        default=2,
        help="Existing process_1.id to populate.",
    )
    parser.add_argument(
        "--process-1-name",
        default="Управление персоналом",
        help="Target process_1.f1_name.",
    )
    return parser.parse_args()


def clean(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def load_structure(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["HR X_v2"]

    level_1 = None
    current_l2 = None
    current_l3 = None

    l2_map = OrderedDict()
    l3_map = OrderedDict()
    l4_items = []

    for row_index in range(1, ws.max_row + 1):
        a, b, c, d = [clean(ws.cell(row_index, col).value) for col in range(1, 5)]
        if not any([a, b, c, d]) or row_index == 3:
            continue

        if a and not b and not c and not d:
            level_1 = a
            current_l2 = None
            current_l3 = None
            continue

        if b:
            current_l2 = b
            current_l3 = None
            l2_map.setdefault(b, {"name": b, "l3": OrderedDict()})
            continue

        if c:
            if current_l2 is None:
                raise ValueError(f"Row {row_index}: L3 without parent L2")
            current_l3 = c
            l2_map[current_l2]["l3"].setdefault(c, {"name": c, "l4": []})
            l3_map[(current_l2, c)] = l2_map[current_l2]["l3"][c]
            continue

        if d:
            if current_l2 is None:
                raise ValueError(f"Row {row_index}: L4 without parent L2")
            if current_l3 is None:
                raise ValueError(f"Row {row_index}: L4 without parent L3")
            l2_map[current_l2]["l3"][current_l3]["l4"].append(d)
            l4_items.append({"l2": current_l2, "l3": current_l3, "l4": d, "source_row": row_index})

    if not level_1:
        raise ValueError("Top-level process not found in workbook")

    structure = {
        "level_1_source": level_1,
        "level_1_target": "Управление персоналом",
        "level_2": [],
    }
    for l2_name, l2_data in l2_map.items():
        l2_entry = {"name": l2_name, "level_3": []}
        for l3_name, l3_data in l2_data["l3"].items():
            l2_entry["level_3"].append({"name": l3_name, "level_4": list(l3_data["l4"])})
        structure["level_2"].append(l2_entry)

    return structure


def sql_quote(value):
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def generate_sql(structure, process_1_id, process_1_name):
    lines = [
        "BEGIN;",
        "",
        "-- Keep existing BUiNU structure intact; only repopulate the existing HR branch.",
        f"UPDATE process_1 SET f1_name = {sql_quote(process_1_name)}, is_active = TRUE, sort = COALESCE(sort, {process_1_id}) WHERE id = {process_1_id};",
        f"DELETE FROM process_2 WHERE process_1_id = {process_1_id};",
        "",
    ]

    lines.append("WITH")
    l2_ctes = []
    l3_ctes = []
    l4_ctes = []

    for l2_sort, l2 in enumerate(structure["level_2"], start=1):
        l2_key = f"l2_{l2_sort}"
        l2_ctes.append(
            f"""{l2_key} AS (
    INSERT INTO process_2 (process_1_id, f2_name, sort, note, is_active)
    VALUES ({process_1_id}, {sql_quote(l2["name"])}, {l2_sort}, NULL, TRUE)
    RETURNING id
)"""
        )

        for l3_sort, l3 in enumerate(l2["level_3"], start=1):
            l3_key = f"l3_{l2_sort}_{l3_sort}"
            l3_ctes.append(
                f"""{l3_key} AS (
    INSERT INTO process_3 (process_2_id, f3_name, sort, note, is_active)
    SELECT id, {sql_quote(l3["name"])}, {l3_sort}, NULL, TRUE
    FROM {l2_key}
    RETURNING id
)"""
            )

            for l4_sort, l4_name in enumerate(l3["level_4"], start=1):
                l4_key = f"l4_{l2_sort}_{l3_sort}_{l4_sort}"
                l4_ctes.append(
                    f"""{l4_key} AS (
    INSERT INTO process_4 (process_3_id, f4_name, sort, note, is_active, executor_id)
    SELECT id, {sql_quote(l4_name)}, {l4_sort}, NULL, TRUE, NULL
    FROM {l3_key}
    RETURNING id
)"""
                )

    all_ctes = l2_ctes + l3_ctes + l4_ctes
    if all_ctes:
        lines.append(",\n".join(all_ctes))
        last_cte = all_ctes[-1].split(" AS ", 1)[0]
        lines.append(f"SELECT COUNT(*) FROM {last_cte};")
    else:
        lines.append("noop AS (SELECT 1)")
        lines.append("SELECT 1;")

    lines.extend(
        [
            "",
            "COMMIT;",
        ]
    )
    return "\n".join(lines) + "\n"


def main():
    args = parse_args()
    structure = load_structure(args.input)

    output_sql = Path(args.output_sql)
    output_json = Path(args.output_json)
    output_sql.parent.mkdir(parents=True, exist_ok=True)
    output_json.parent.mkdir(parents=True, exist_ok=True)

    sql = generate_sql(structure, args.process_1_id, args.process_1_name)
    output_sql.write_text(sql, encoding="utf-8")
    output_json.write_text(json.dumps(structure, ensure_ascii=False, indent=2), encoding="utf-8")

    stats = {
        "level_1_target": args.process_1_name,
        "level_2_count": len(structure["level_2"]),
        "level_3_count": sum(len(l2["level_3"]) for l2 in structure["level_2"]),
        "level_4_count": sum(len(l3["level_4"]) for l2 in structure["level_2"] for l3 in l2["level_3"]),
        "output_sql": str(output_sql),
        "output_json": str(output_json),
    }
    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
